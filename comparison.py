import requests
from azure.identity import ClientSecretCredential
import csv
import openpyxl
from datetime import datetime
import os
from config import TENANT_ID, CLIENT_ID, CLIENT_SECRET, WORKSPACE_ID, DATASET_ID, table_name, file_path, LOG_FILE, SHEET_NAME, SCOPE
import logging

logging.basicConfig(
    level=logging.INFO,  
    format="%(asctime)s - %(levelname)s - %(message)s"
)

logger = logging.getLogger(__name__)

logging.getLogger("azure").setLevel(logging.WARNING)
logging.getLogger("azure.core.pipeline.policies.http_logging_policy").setLevel(logging.WARNING)

# Configuration
credential = ClientSecretCredential(
    tenant_id=TENANT_ID,
    client_id=CLIENT_ID,
    client_secret=CLIENT_SECRET
)

token = credential.get_token(
    SCOPE
).token

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

# API endpoint
url = f"https://api.powerbi.com/v1.0/myorg/groups/{WORKSPACE_ID}/datasets/{DATASET_ID}/executeQueries"


# DAX quert to fetch row count from Dataset
def run_dax(query):
    payload = {"queries": [{"query": query}]}
    res = requests.post(url, json=payload, headers=headers)
    return res.json()

dax_rowcount = f"""
EVALUATE
ROW("RowCount", COUNTROWS('{table_name}'))
"""

dataset_result = run_dax(dax_rowcount)
# print(dataset_result)

try:
    dataset_row_count = list(
        dataset_result["results"][0]["tables"][0]["rows"][0].values()
    )[0]

    logger.info(f"Dataset Row Count: {dataset_row_count}")

except Exception as e:
    dataset_row_count = None
    logger.error("Could not extract dataset row count: %s", e)

# Fetching row count from source file
def csv_row_count(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError("File not found: " + file_path)

    count = 0
    with open(file_path, "r", encoding="utf-8") as f:
        for _ in f:
            count += 1
    return count - 1 

csv_count = csv_row_count(file_path)
logger.info(f"CSV Rows (excluding header): {csv_count}")


# Comparing data : Matched or Mismatched
if dataset_row_count == csv_count:
    logger.info(f"MATCH: Dataset rows = {dataset_row_count}, CSV rows = {csv_count}")
else:
    logger.info(f"MISMATCH: Dataset rows = {dataset_row_count}, CSV rows = {csv_count}")


# Create or update the excel with comparison
def log_to_excel(dataset_count, csv_count):
    status = "MATCH" if dataset_count == csv_count else "MISMATCH"

    if not os.path.exists(LOG_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = SHEET_NAME
        sheet.append(["Run Date","Dataset ID", "Dataset Count", "CSV Count", "Status"])
        wb.save(LOG_FILE)

    wb = openpyxl.load_workbook(LOG_FILE)
    sheet = wb[SHEET_NAME]
    next_row = sheet.max_row + 1

    sheet.cell(row=next_row, column=1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.cell(row=next_row, column=2).value = DATASET_ID
    sheet.cell(row=next_row, column=3).value = dataset_count
    sheet.cell(row=next_row, column=4).value = csv_count
    sheet.cell(row=next_row, column=5).value = status

    wb.save(LOG_FILE)
    logger.info(f"Log saved to {LOG_FILE}")

log_to_excel(dataset_row_count, csv_count)